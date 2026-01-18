"""
Excel Episode Logger für Trajektorie-Analyse nach jeder Episode.

Speichert für jede Episode:
- Datum und Zeit
- Controller Parameter
- Trajektorie-Daten pro Phase (Anzahl Wegpunkte, Zeitdauer, Modifikatoren)
"""

import os
from datetime import datetime
from pathlib import Path
import numpy as np

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelEpisodeLogger:
    """
    Logger für Episode-Daten in Excel-Format.
    
    Speichert nach jeder Episode:
    - Zeitstempel (Datum, Zeit)
    - Controller Parameter (trajectory_resolution, air_speed_multiplier, etc.)
    - Phase-Daten (10 Phasen: Wegpunkte, Zeit, Modifikatoren)
    """
    
    # 10 Phasen des Pick&Place Controllers
    PHASES = [
        "GRIP_OPEN",              # 0
        "MOVE_DOWN_CRITICAL",     # 1
        "GRIP_CLOSE",             # 2
        "MOVE_UP",                # 3
        "MOVE_TO_STACK",          # 4
        "MOVE_DOWN_CRITICAL_STK", # 5
        "WAIT",                   # 6
        "GRIP_OPEN_STK",          # 7
        "MOVE_UP_STK",            # 8
        "MOVE_AWAY",              # 9
    ]
    
    def __init__(self, output_dir: str = None, filename: str = None):
        """
        Initialisiert den Excel Logger.
        
        Args:
            output_dir: Verzeichnis für Excel-Datei (default: ./logs)
            filename: Name der Excel-Datei (default: episode_tracking.xlsx)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl nicht installiert. Installiere mit: pip install openpyxl")
        
        self.output_dir = Path(output_dir or "./logs")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.filename = filename or "episode_tracking.xlsx"
        self.filepath = self.output_dir / self.filename
        
        # Initialisiere oder lade Workbook
        self._initialize_workbook()
    
    def _initialize_workbook(self):
        """Erstellt oder lädt existierendes Workbook mit Header."""
        if self.filepath.exists():
            self.workbook = load_workbook(self.filepath)
            self.ws = self.workbook.active
        else:
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = "Episodes"
            self._create_header()
    
    def _create_header(self):
        """Erstellt Header-Zeile mit Spalten."""
        headers = [
            "Episode ID",
            "Datum",
            "Zeit",
            "trajectory_resolution",
            "air_speed_multiplier",
            "height_adaptive_speed",
            "critical_height_threshold",
            "critical_speed_factor",
            "Gesamte Timesteps",
            "Gesamtzeit (s)",
        ]
        
        # Phase-Header: Für jede Phase 3 Spalten (Wegpunkte, Zeit, Modifikator)
        for phase_name in self.PHASES:
            headers.extend([
                f"{phase_name}_Wegpunkte",
                f"{phase_name}_Zeit (s)",
                f"{phase_name}_Modifikator",
            ])
        
        # Zusätzliche Spalten
        headers.extend([
            "Validierung erfolgreich",
            "Notizen",
        ])
        
        # Schreibe Header
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Setze Spaltenbreiten
        self.ws.column_dimensions['A'].width = 12
        self.ws.column_dimensions['B'].width = 12
        self.ws.column_dimensions['C'].width = 12
        for col_idx in range(4, 10):
            self.ws.column_dimensions[chr(64 + col_idx)].width = 10
        for col_idx in range(10, 10 + len(self.PHASES) * 3):
            self.ws.column_dimensions[self.get_column_letter(col_idx)].width = 12
        self.ws.column_dimensions[self.get_column_letter(10 + len(self.PHASES) * 3)].width = 15
        self.ws.column_dimensions[self.get_column_letter(10 + len(self.PHASES) * 3 + 1)].width = 20
        
        # Friere Header ein
        self.ws.freeze_panes = "A2"
    
    @staticmethod
    def get_column_letter(col_idx):
        """Konvertiert Spalten-Index zu Excel-Buchstabe."""
        result = ""
        col_idx -= 1
        while col_idx >= 0:
            result = chr(col_idx % 26 + 65) + result
            col_idx = col_idx // 26 - 1
        return result
    
    def log_episode(
        self,
        episode_id: int,
        controller_params: dict,
        phase_data: dict,
        total_timesteps: int,
        total_time: float,
        validation_success: bool = True,
        notes: str = None,
    ):
        """
        Loggt eine abgeschlossene Episode in Excel.
        
        Args:
            episode_id: Eindeutige Episode-ID
            controller_params: Dict mit Controller-Parametern:
                {
                    "trajectory_resolution": float,
                    "air_speed_multiplier": float,
                    "height_adaptive_speed": bool,
                    "critical_height_threshold": float,
                    "critical_speed_factor": float,
                }
            phase_data: Dict mit Phase-Daten:
                {
                    phase_idx: {
                        "waypoints": int,  # Anzahl Trajektorienpunkte
                        "time": float,     # Zeit in Sekunden
                        "modifier": float, # dt-Modifikator (z.B. 1.0, 4.0)
                    },
                    ...
                }
            total_timesteps: Gesamtzahl Timesteps für die Episode
            total_time: Gesamtzeit der Episode in Sekunden
            validation_success: Ob Validierung erfolgreich war
            notes: Optionale Notizen
        """
        # Finde nächste freie Zeile
        next_row = self.ws.max_row + 1
        
        now = datetime.now()
        
        col_idx = 1
        
        # Grundlegende Spalten
        self.ws.cell(row=next_row, column=col_idx).value = episode_id
        col_idx += 1
        
        self.ws.cell(row=next_row, column=col_idx).value = now.strftime("%d.%m.%Y")
        col_idx += 1
        
        self.ws.cell(row=next_row, column=col_idx).value = now.strftime("%H:%M:%S")
        col_idx += 1
        
        # Controller Parameter
        self.ws.cell(row=next_row, column=col_idx).value = controller_params.get("trajectory_resolution", 1.0)
        col_idx += 1
        
        self.ws.cell(row=next_row, column=col_idx).value = controller_params.get("air_speed_multiplier", 1.0)
        col_idx += 1
        
        height_adaptive = controller_params.get("height_adaptive_speed", False)
        self.ws.cell(row=next_row, column=col_idx).value = "JA" if height_adaptive else "NEIN"
        col_idx += 1
        
        self.ws.cell(row=next_row, column=col_idx).value = controller_params.get("critical_height_threshold", 0.15)
        col_idx += 1
        
        self.ws.cell(row=next_row, column=col_idx).value = controller_params.get("critical_speed_factor", 0.25)
        col_idx += 1
        
        # Gesamtdaten
        self.ws.cell(row=next_row, column=col_idx).value = total_timesteps
        col_idx += 1
        
        self.ws.cell(row=next_row, column=col_idx).value = round(total_time, 4)
        col_idx += 1
        
        # Phase-Daten (10 Phasen)
        for phase_idx in range(len(self.PHASES)):
            if phase_idx in phase_data:
                phase_info = phase_data[phase_idx]
                
                # Wegpunkte
                self.ws.cell(row=next_row, column=col_idx).value = phase_info.get("waypoints", 0)
                col_idx += 1
                
                # Zeit
                self.ws.cell(row=next_row, column=col_idx).value = round(phase_info.get("time", 0.0), 4)
                col_idx += 1
                
                # Modifikator
                self.ws.cell(row=next_row, column=col_idx).value = round(phase_info.get("modifier", 1.0), 4)
                col_idx += 1
            else:
                # Leere Phasen-Daten
                self.ws.cell(row=next_row, column=col_idx).value = 0
                col_idx += 1
                self.ws.cell(row=next_row, column=col_idx).value = 0.0
                col_idx += 1
                self.ws.cell(row=next_row, column=col_idx).value = 1.0
                col_idx += 1
        
        # Validierung
        self.ws.cell(row=next_row, column=col_idx).value = "✓ JA" if validation_success else "✗ NEIN"
        col_idx += 1
        
        # Notizen
        self.ws.cell(row=next_row, column=col_idx).value = notes or ""
        
        # Formatierung
        self._format_row(next_row, validation_success)
        
        # Speichere Workbook
        self.workbook.save(self.filepath)
    
    def _format_row(self, row_idx: int, success: bool = True):
        """Formatiert eine Zeile mit Farben je nach Erfolg."""
        fill_color = "D4EDDA" if success else "F8D7DA"  # Grün oder Rot
        
        for col_idx in range(1, self.ws.max_column + 1):
            cell = self.ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Zahlen-Formatierung für numerische Spalten
            if col_idx >= 4 and col_idx < 10 + len(self.PHASES) * 3:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0000'


def extract_phase_data_from_controller(controller) -> tuple:
    """
    Extrahiert Phase-Daten aus dem Controller nach einer Episode.
    
    Versucht, Daten aus Controller-Attributen zu extrahieren.
    
    Returns:
        tuple: (total_timesteps, total_time, phase_data_dict)
        
    phase_data_dict = {
        phase_idx: {
            "waypoints": int,
            "time": float,
            "modifier": float,
        },
        ...
    }
    """
    total_timesteps = 0
    total_time = 0.0
    phase_data = {}
    
    # Versuche Controller-Attribute zu extrahieren
    try:
        # Falls Controller Phase-Tracking hat
        if hasattr(controller, '_phase_timesteps'):
            for phase_idx, timesteps in controller._phase_timesteps.items():
                phase_data[phase_idx] = {
                    "waypoints": timesteps,
                    "time": timesteps * (1.0 / 60.0),  # Annahme: 60 Hz Simulation
                    "modifier": 1.0,
                }
                total_timesteps += timesteps
                total_time += timesteps * (1.0 / 60.0)
        
        # Falls Controller Gesamtdaten hat
        if hasattr(controller, '_total_timesteps'):
            total_timesteps = controller._total_timesteps
        
        if hasattr(controller, '_total_time'):
            total_time = controller._total_time
    
    except Exception as e:
        print(f"Warnung: Konnte Phase-Daten nicht extrahieren: {e}")
    
    return total_timesteps, total_time, phase_data
