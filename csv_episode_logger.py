"""
CSV Episode Logger für Trajektorie-Analyse nach jeder Episode.

Speichert Daten im Matrix-Format:
- Reihen: Variablen (Datum, Zeit, Parameter, Phase-Metriken)
- Spalten: Einzelne Runs/Episodes

Verwendet CSV-Format (kompatibel mit Excel, keine Abhängigkeiten).
"""

import csv
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class CSVEpisodeLogger:
    """
    Logger für Episode-Daten im CSV-Matrix-Format.
    
    Format: 
    - Reihen = Variablen (Datum, Zeit, Parameter, Phase-Daten)
    - Spalten = Einzelne Episodes/Runs
    
    Speichert alle Episode-Daten im Memory und schreibt Matrix am Ende.
    """
    
    # 10 Phasen des Pick&Place Controllers mit Beschreibung
    PHASES = [
        ("Phase 0", "GRIP_OPEN"),           # 0 - Move EE above cube [AIR]
        ("Phase 1", "MOVE_DOWN"),  # 1 - Lower EE to cube [CRITICAL]
        ("Phase 2", "GRIP_CLOSE"),          # 2 - Wait for settle [WAIT]
        ("Phase 3", "MOVE_UP"),             # 3 - Close gripper [GRIP]
        ("Phase 4", "MOVE_TO_STACK"),       # 4 - Lift EE up [AIR]
        ("Phase 5", "MOVE_DOWN"), # 5 - Move to target XY [AIR]
        ("Phase 6", "WAIT"),                # 6 - Lower to place [CRITICAL]
        ("Phase 7", "GRIP_OPEN"),       # 7 - Open gripper [RELEASE]
        ("Phase 8", "MOVE_UP"),         # 8 - Lift EE up [AIR]
        ("Phase 9", "MOVE_AWAY"),           # 9 - Return to start [AIR]
    ]

        
    
    def __init__(self, output_dir: str = None, filename: str = None):
        """
        Initialisiert den CSV Logger im Matrix-Format.
        
        Args:
            output_dir: Verzeichnis für CSV-Datei (default: ./logs)
            filename: Name der CSV-Datei (default: episode_tracking.csv)
        """
        # output_dir kommt von logger.dataset_path, das endet mit 000000/
        # Gehe deshalb ein Verzeichnis nach oben
        self.output_dir = Path(output_dir).parent
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.filename = filename or "episode_tracking.csv"
        self.filepath = self.output_dir / self.filename
        
        # Sammele alle Episode-Daten im Memory (Spalten = Episodes, Reihen = Variablen)
        self.episodes_data = {}
    
    def log_episode(
        self,
        episode_seed: int,
        controller_params: dict,
        phase_data: dict,
        total_timesteps: int,
        total_time: float,
        validation_success: bool = True,
        notes: str = None,
    ):
        """
        Sammelt Episode-Daten für Matrix-Format (nicht sofort schreiben).
        
        Args:
            episode_seed: Eindeutige Episode-ID oder Seed
            controller_params: Dict mit Controller-Parametern
            phase_data: Dict mit Phase-Daten pro Phase-Index
            total_timesteps: Gesamtzahl Timesteps für die Episode
            total_time: Gesamtzeit der Episode in Sekunden
            validation_success: Ob Validierung erfolgreich war
            notes: Optionale Notizen
        """
        now = datetime.now()
        
        # Erstelle Spalte für diese Episode
        run_key = f"Run {len(self.episodes_data)}"
        
        column_data = {}
        
        # === METADATEN ===
        column_data["Datum"] = now.strftime("%d.%m.%Y")
        column_data["Zeit"] = now.strftime("%H:%M:%S")
        
        # === CONTROLLER PARAMETER ===
        column_data["trajectory_resolution"] = str(controller_params.get("trajectory_resolution", 1.0)).replace('.', ',')
        column_data["air_speed_multiplier"] = str(controller_params.get("air_speed_multiplier", 1.0)).replace('.', ',')
        column_data["height_adaptive_speed"] = "JA" if controller_params.get("height_adaptive_speed", False) else "NEIN"
        column_data["critical_height_threshold"] = str(controller_params.get("critical_height_threshold", 0.15)).replace('.', ',')
        column_data["critical_speed_factor"] = str(controller_params.get("critical_speed_factor", 0.25)).replace('.', ',')
        
        # === GESAMTE EPISODE-METRIKEN ===
        column_data["Gesamte Timesteps"] = str(total_timesteps)
        column_data["Gesamtzeit (s)"] = str(round(total_time, 4)).replace('.', ',')
        
        # === PHASE-DATEN ===
        for phase_idx, (phase_label, phase_name) in enumerate(self.PHASES):
            if phase_idx in phase_data:
                phase_info = phase_data[phase_idx]
                column_data[f"{phase_label}: {phase_name} - Wegpunkte"] = str(phase_info.get("waypoints", 0))
                column_data[f"{phase_label}: {phase_name} - Zeit (s)"] = str(round(phase_info.get("time", 0.0), 4)).replace('.', ',')
            else:
                column_data[f"{phase_label}: {phase_name} - Wegpunkte"] = "0"
                column_data[f"{phase_label}: {phase_name} - Zeit (s)"] = "0,0"
        
        # === VALIDIERUNG UND NOTIZEN ===
        column_data["Validierung erfolgreich"] = "JA" if validation_success else "NEIN"
        notes_safe = (notes or "").replace(';', ',')
        column_data["Notizen"] = notes_safe
        
        # Speichere Episode-Spalte
        self.episodes_data[run_key] = column_data
        
        print(f"✓ Episode {len(self.episodes_data)} gesammelt: {run_key}")
    
    def save_matrix(self):
        """
        Schreibt alle gesammelten Episode-Daten ins Excel/CSV.
        
        Format:
        - Reihen = Variablen/Metriken
        - Spalten = Episodes (Run 0, Run 1, Run 2, ...)
        
        Falls Excel-Datei bereits existiert, fügt neue Daten in erste freie Spalte ein.
        Sonst erstellt neue Datei.
        """
        if not self.episodes_data:
            print("⚠ Keine Episode-Daten zum Speichern vorhanden")
            return
        
        try:
            # Sortiere Reihen in gewünschte Reihenfolge
            rows_order = [
                "Datum",
                "Zeit",
                "trajectory_resolution",
                "air_speed_multiplier",
                "height_adaptive_speed",
                "critical_height_threshold",
                "critical_speed_factor",
                "Gesamte Timesteps",
                "Gesamtzeit (s)",
            ]
            
            # Füge Phase-Daten in Reihenfolge hinzu
            for phase_idx, (phase_label, phase_name) in enumerate(self.PHASES):
                rows_order.append(f"{phase_label}: {phase_name} - Wegpunkte")
                rows_order.append(f"{phase_label}: {phase_name} - Zeit (s)")
            
            # Füge restliche Reihen hinzu
            rows_order.append("Validierung erfolgreich")
            rows_order.append("Notizen")
            
            # Versuche Excel-Datei zu öffnen (falls openpyxl verfügbar)
            xlsx_path = self.output_dir / (self.filename.replace('.csv', '.xlsx'))
            
            if HAS_OPENPYXL and xlsx_path.exists():
                self._append_to_excel(xlsx_path, rows_order)
            elif HAS_OPENPYXL:
                self._create_new_excel(xlsx_path, rows_order)
            else:
                # Fallback auf CSV wenn openpyxl nicht verfügbar
                self._create_csv(rows_order)
                
        except Exception as e:
            print(f"✗ FEHLER beim Speichern: {e}")
    
    def _create_new_excel(self, xlsx_path, rows_order):
        """Erstellt neue Excel-Datei mit Daten."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Header-Zeile
        column_keys = list(self.episodes_data.keys())
        header = ["Variable"] + column_keys
        ws.append(header)
        
        # Daten-Zeilen
        for row_name in rows_order:
            row = [row_name]
            for run_key in column_keys:
                value = self.episodes_data[run_key].get(row_name, "")
                row.append(value)
            ws.append(row)
        
        wb.save(xlsx_path)
        print(f"✓ Neue Excel-Datei erstellt: {xlsx_path}")
        print(f"  Format: {len(rows_order)} Reihen × {len(column_keys)+1} Spalten")
    
    def _append_to_excel(self, xlsx_path, rows_order):
        """Fügt neue Daten in erste freie Spalte existierender Excel an."""
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        # Finde erste freie Spalte (nach der "Variable" Spalte und existierenden Daten)
        next_col = ws.max_column + 1
        
        # Header: Setze neuen Run-Namen
        column_keys = list(self.episodes_data.keys())
        run_name = column_keys[0]  # z.B. "Run 0", "Run 1", etc.
        ws.cell(row=1, column=next_col, value=run_name)
        
        # Daten-Zeilen: Füge Werte in neue Spalte ein
        for row_idx, row_name in enumerate(rows_order, start=2):
            # Finde oder erstelle Reihe für diese Variable
            row_found = False
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == row_name:
                    ws.cell(row=r, column=next_col, value=self.episodes_data[run_name].get(row_name, ""))
                    row_found = True
                    break
            
            # Falls Reihe nicht existiert, erstelle sie
            if not row_found:
                new_row = ws.max_row + 1
                ws.cell(row=new_row, column=1, value=row_name)
                ws.cell(row=new_row, column=next_col, value=self.episodes_data[run_name].get(row_name, ""))
        
        wb.save(xlsx_path)
        print(f"✓ Daten angefügt zu Excel: {xlsx_path}")
        print(f"  Neue Spalte: {next_col} ({run_name})")
    
    def _create_csv(self, rows_order):
        """Fallback: Fügt CSV-Daten zur existierenden Datei oder erstellt neu."""
        csv_path = self.filepath
        column_keys = list(self.episodes_data.keys())
        run_name = column_keys[0]
        
        # Prüfe ob CSV existiert
        if csv_path.exists():
            # Lese existierende Datei
            existing_data = {}
            existing_columns = []  # Behalte Spalten-Reihenfolge
            
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f, delimiter=';')
                # Speichere Header-Reihenfolge
                if reader.fieldnames:
                    existing_columns = [col for col in reader.fieldnames if col != 'Variable']
                
                for row in reader:
                    var_name = row.pop(None) if None in row else row.pop('Variable', '')
                    existing_data[var_name] = row
            
            # Überprüfe ob run_name bereits existiert und generiere neuen Namen falls nötig
            if run_name in existing_columns:
                # Generiere eindeutigen Namen
                counter = 0
                while f"{run_name[:-1]}{counter}" in existing_columns:
                    counter += 1
                run_name = f"{run_name[:-1]}{counter}"
                print(f"⚠ Spalte 'Run X' existiert bereits, verwende stattdessen: {run_name}")
            
            # Füge neue Spalte hinzu (als letzte Spalte)
            for row_name in rows_order:
                if row_name not in existing_data:
                    existing_data[row_name] = {}
                existing_data[row_name][run_name] = self.episodes_data[column_keys[0]].get(row_name, "")
            
            # Beibehalte Spalten-Reihenfolge und füge neue Spalte am Ende an
            all_columns = existing_columns + [run_name]
            
            # Schreibe CSV neu mit neuer Spalte
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                
                # Header
                header = ["Variable"] + all_columns
                writer.writerow(header)
                
                # Daten
                for var_name in rows_order:
                    row = [var_name]
                    for col in all_columns:
                        value = existing_data.get(var_name, {}).get(col, "")
                        row.append(value)
                    writer.writerow(row)
            
            print(f"✓ CSV aktualisiert (neue Spalte angefügt): {csv_path}")
            print(f"  Neue Spalte: {run_name}")
        else:
            # Erstelle neue CSV
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                
                # Header-Zeile
                header = ["Variable"] + column_keys
                writer.writerow(header)
                
                # Daten-Zeilen
                for row_name in rows_order:
                    row = [row_name]
                    for run_key in column_keys:
                        value = self.episodes_data[run_key].get(row_name, "")
                        row.append(value)
                    writer.writerow(row)
            
            print(f"✓ CSV erstellt (openpyxl nicht verfügbar): {csv_path}")
        
        print(f"  Format: {len(rows_order)} Reihen")
